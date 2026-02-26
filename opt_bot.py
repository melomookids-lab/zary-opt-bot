"""
ZARY & CO OPT Bot — Production Version
Особенности: асинхронная БД, автоотчеты в конце месяца, полная статистика
"""

import os
import re
import asyncio
import logging
import sqlite3
import aiosqlite
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, Any, List
from calendar import monthrange

from aiohttp import web
from apscheduler.schedulers.asyncio import AsyncIOScheduler

from aiogram import Bot, Dispatcher, F
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart, Command
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
)
from aiogram.client.default import DefaultBotProperties
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types.input_file import FSInputFile
from aiogram.exceptions import TelegramAPIError

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# =========================
# CONFIGURATION
# =========================
class Config:
    BOT_TOKEN = (os.getenv("BOT_TOKEN") or "").strip()
    MANAGER_ID_RAW = (os.getenv("MANAGER_ID") or "").strip()
    CHANNEL = (os.getenv("CHANNEL") or "zaryco_official").strip().lstrip("@")
    PHONE = (os.getenv("PHONE") or "+998771202255").strip()
    PORT = int((os.getenv("PORT") or "10000").strip())
    
    DB_PATH = "leads.sqlite3"
    EXPORTS_DIR = Path("exports")
    BACKUP_DIR = Path("backups")
    REPORTS_DIR = Path("reports")  # Для месячных отчетов
    
    # Настройки отчетов
    MAX_EXPORT_AGE_DAYS = 7
    BACKUP_KEEP_COUNT = 5
    
    # Валидация
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN не указан!")
    if not MANAGER_ID_RAW.isdigit():
        raise RuntimeError("MANAGER_ID должен быть числом!")
    
    MANAGER_ID = int(MANAGER_ID_RAW)

# =========================
# LOGGING
# =========================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =========================
# ASYNC DATABASE
# =========================
class Database:
    def __init__(self, db_path: str):
        self.db_path = db_path
        self._pool: Optional[aiosqlite.Connection] = None
    
    async def connect(self):
        self._pool = await aiosqlite.connect(self.db_path)
        self._pool.row_factory = aiosqlite.Row
        await self._pool.execute("PRAGMA foreign_keys = ON")
        await self._pool.execute("PRAGMA journal_mode = WAL")
        await self.init_tables()
        logger.info("Database connected")
    
    async def close(self):
        if self._pool:
            await self._pool.close()
    
    async def init_tables(self):
        await self._pool.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                lang TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                last_activity TEXT DEFAULT CURRENT_TIMESTAMP
            );
            
            CREATE TABLE IF NOT EXISTS leads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                user_id INTEGER NOT NULL,
                username TEXT,
                full_name TEXT,
                lang TEXT NOT NULL,
                role TEXT NOT NULL,
                product TEXT NOT NULL,
                qty TEXT NOT NULL,
                city TEXT NOT NULL,
                phone TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'new',
                manager_notified INTEGER DEFAULT 0,
                notes TEXT
            );
            
            CREATE INDEX IF NOT EXISTS idx_leads_user_id ON leads(user_id);
            CREATE INDEX IF NOT EXISTS idx_leads_status ON leads(status);
            CREATE INDEX IF NOT EXISTS idx_leads_created ON leads(created_at);
            
            CREATE TABLE IF NOT EXISTS activity_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
                user_id INTEGER,
                action TEXT,
                details TEXT
            );
            
            CREATE TABLE IF NOT EXISTS monthly_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                sent_at TEXT NOT NULL,
                filename TEXT NOT NULL,
                total_leads INTEGER NOT NULL,
                status TEXT DEFAULT 'sent'
            );
        """)
        await self._pool.commit()
    
    async def get_lang(self, user_id: int) -> Optional[str]:
        async with self._pool.execute(
            "SELECT lang FROM users WHERE user_id = ?", (user_id,)
        ) as cursor:
            row = await cursor.fetchone()
            return row[0] if row else None
    
    async def set_lang(self, user_id: int, lang: str):
        await self._pool.execute("""
            INSERT INTO users(user_id, lang, last_activity) 
            VALUES(?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(user_id) DO UPDATE 
            SET lang = excluded.lang, last_activity = CURRENT_TIMESTAMP
        """, (user_id, lang))
        await self._pool.commit()
    
    async def add_lead(self, lead: Dict[str, Any]) -> int:
        cursor = await self._pool.execute("""
            INSERT INTO leads (
                created_at, user_id, username, full_name, lang, 
                role, product, qty, city, phone, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            lead["created_at"], lead["user_id"], lead.get("username"),
            lead.get("full_name"), lead["lang"], lead["role"],
            lead["product"], lead["qty"], lead["city"], lead["phone"], "new"
        ))
        await self._pool.commit()
        return cursor.lastrowid
    
    async def get_last_leads(self, limit: int = 20) -> List[aiosqlite.Row]:
        async with self._pool.execute("""
            SELECT * FROM leads 
            ORDER BY id DESC 
            LIMIT ?
        """, (limit,)) as cursor:
            return await cursor.fetchall()
    
    async def get_all_leads(self) -> List[aiosqlite.Row]:
        async with self._pool.execute("""
            SELECT * FROM leads ORDER BY id DESC
        """,) as cursor:
            return await cursor.fetchall()
    
    async def get_leads_by_date_range(self, start_date: str, end_date: str) -> List[aiosqlite.Row]:
        """Получить заказы за период (YYYY-MM-DD)"""
        async with self._pool.execute("""
            SELECT * FROM leads 
            WHERE created_at >= ? AND created_at <= ?
            ORDER BY id DESC
        """, (start_date, end_date)) as cursor:
            return await cursor.fetchall()
    
    async def update_status(self, lead_id: int, status: str) -> bool:
        cursor = await self._pool.execute("""
            UPDATE leads SET status = ? WHERE id = ?
        """, (status, lead_id))
        await self._pool.commit()
        return cursor.rowcount > 0
    
    async def update_notification_status(self, lead_id: int, notified: bool):
        await self._pool.execute("""
            UPDATE leads SET manager_notified = ? WHERE id = ?
        """, (1 if notified else 0, lead_id))
        await self._pool.commit()
    
    async def log_activity(self, user_id: int, action: str, details: str = ""):
        await self._pool.execute("""
            INSERT INTO activity_log (user_id, action, details) 
            VALUES (?, ?, ?)
        """, (user_id, action, details))
        await self._pool.commit()
    
    async def get_stats(self) -> Dict[str, int]:
        async with self._pool.execute("""
            SELECT 
                COUNT(*) as total_leads,
                SUM(CASE WHEN status = 'new' THEN 1 ELSE 0 END) as new_leads,
                COUNT(DISTINCT user_id) as unique_users
            FROM leads
        """) as cursor:
            row = await cursor.fetchone()
            return dict(row) if row else {}
    
    async def get_monthly_stats(self, year: int, month: int) -> Dict[str, Any]:
        """Статистика за месяц"""
        start = f"{year}-{month:02d}-01"
        last_day = monthrange(year, month)[1]
        end = f"{year}-{month:02d}-{last_day} 23:59:59"
        
        async with self._pool.execute("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN status = 'new' THEN 1 ELSE 0 END) as new_count,
                SUM(CASE WHEN status = 'work' THEN 1 ELSE 0 END) as work_count,
                SUM(CASE WHEN status = 'paid' THEN 1 ELSE 0 END) as paid_count,
                SUM(CASE WHEN status = 'shipped' THEN 1 ELSE 0 END) as shipped_count,
                SUM(CASE WHEN status = 'closed' THEN 1 ELSE 0 END) as closed_count,
                COUNT(DISTINCT user_id) as unique_clients
            FROM leads 
            WHERE created_at >= ? AND created_at <= ?
        """, (start, end)) as cursor:
            row = await cursor.fetchone()
            return {
                "period": f"{month:02d}.{year}",
                "start": start,
                "end": end[:10],
                **dict(row)
            }
    
    async def mark_report_sent(self, year: int, month: int, filename: str, total_leads: int):
        """Отметить что отчет отправлен"""
        await self._pool.execute("""
            INSERT INTO monthly_reports (year, month, sent_at, filename, total_leads)
            VALUES (?, ?, CURRENT_TIMESTAMP, ?, ?)
        """, (year, month, filename, total_leads))
        await self._pool.commit()
    
    async def is_report_sent(self, year: int, month: int) -> bool:
        """Проверить отправлен ли отчет за месяц"""
        async with self._pool.execute("""
            SELECT 1 FROM monthly_reports 
            WHERE year = ? AND month = ? AND status = 'sent'
        """, (year, month)) as cursor:
            return await cursor.fetchone() is not None

db = Database(Config.DB_PATH)

# =========================
# BOT SETUP
# =========================
bot = Bot(
    Config.BOT_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)
dp = Dispatcher(storage=MemoryStorage())

# =========================
# TEXTS
# =========================
def t(key: str, lang: str = "ru", **kwargs) -> str:
    texts = {
        "ru": {
            "welcome": "🤝 <b>ZARY & CO — ОПТОВЫЙ ОТДЕЛ</b>\n\n"
                      "Работаем с магазинами, бутиками и маркетплейсами.\n"
                      "• Национальный бренд\n"
                      "• Стабильные поставки\n"
                      "• Высокая маржинальность\n\n"
                      "👇 Выберите действие:",
            
            "menu_hint": "📍 Главное меню",
            
            "manager": f"📞 <b>Менеджер оптового отдела</b>\n\n"
                      f"Телефон: <code>{Config.PHONE}</code>\n"
                      f"Режим работы: Пн-Пт 9:00-18:00",
            
            "channel": f"📣 <b>Наш канал с коллекциями</b>\n\n"
                      f"👉 https://t.me/{Config.CHANNEL}",
            
            "catalog": f"📸 <b>Актуальный каталог</b>\n\n"
                      f"Смотрите в нашем канале:\n"
                      f"👉 https://t.me/{Config.CHANNEL}",
            
            "terms": "🧾 <b>Условия сотрудничества</b>\n\n"
                    "✅ Форма работы: предзаказ / наличие\n"
                    "✅ Минимальный заказ: от 20 единиц\n"
                    "✅ Доставка: по всему Узбекистану\n"
                    "✅ Оплата: перечисление / наличные",
            
            "why": "⭐ <b>Почему выбирают нас</b>\n\n"
                  "🏆 Опыт: 5+ лет на рынке\n"
                  "🏭 Производство: собственное в Ташкенте\n"
                  "📦 Ассортимент: 500+ моделей\n"
                  "🚚 Логистика: 2-3 дня по всей стране",
            
            "min_order": "📦 <b>Минимальный заказ</b>\n\n"
                        "• Опт: от 20 единиц\n"
                        "• Крупный опт: от 100 единиц\n"
                        "• Дропшиппинг: индивидуально\n\n"
                        "Хотите персональный расчёт?",
            
            "min_cta": "✅ Оставить заявку",
            
            "form_role": "👤 <b>Кто вы?</b>\n\nВыберите тип бизнеса:",
            "form_product": "👕 <b>Что хотите заказать?</b>\n\n"
                           "Выберите или напишите свой вариант:",
            "form_qty": "📊 <b>Объём заказа?</b>",
            "form_city": "📍 <b>Город доставки?</b>",
            "form_phone": "📱 <b>Контактный телефон</b>\n\n"
                         "Нажмите кнопку «📲 Отправить контакт» или введите вручную:",
            
            "bad_phone": "❌ <b>Некорректный номер</b>\n\n"
                        "Пример: +998901234567",
            
            "thanks": lambda lead_id: 
                     f"✅ <b>Заявка #{lead_id} принята!</b>\n\n"
                     f"Менеджер свяжется с вами в течение 15 минут.\n\n"
                     f"📣 https://t.me/{Config.CHANNEL}\n"
                     f"📞 {Config.PHONE}",
            
            "cancelled": "❌ Отменено. Возвращаюсь в меню...",
            
            "admin_only": "⛔ Только для администратора.",
            "admin_menu": "🛠 <b>Панель управления</b>",
            "admin_last": "📋 <b>Последние заявки</b>\n\n",
            "admin_empty": "📝 Пока нет заявок.",
            "admin_export_ok": "✅ Excel сформирован.",
            "admin_export_fail": "❌ Ошибка при создании файла.",
            "admin_status_updated": "✅ Статус обновлён.",
            "admin_status_bad": "❌ Неверная команда.\n\n"
                               "Используйте: /status ID статус\n\n"
                               "Статусы: new, work, paid, shipped, closed",
            
            "monthly_report_subject": lambda p: f"📊 Отчет за {p}",
            "monthly_report_intro": lambda s: 
                f"<b>📊 МЕСЯЧНЫЙ ОТЧЕТ — {s['period']}</b>\n\n"
                f"📅 Период: {s['start']} — {s['end']}\n"
                f"📋 Всего заявок: <b>{s['total']}</b>\n"
                f"👥 Уникальных клиентов: <b>{s['unique_clients']}</b>\n\n"
                f"<b>Статусы:</b>\n"
                f"🆕 Новые: {s['new_count']}\n"
                f"🔧 В работе: {s['work_count']}\n"
                f"💰 Оплачено: {s['paid_count']}\n"
                f"🚚 Отправлено: {s['shipped_count']}\n"
                f"✅ Закрыто: {s['closed_count']}",
            
            "error": "⚠️ Ошибка. Попробуйте позже.",
            "unknown": "🤔 Используйте меню ниже или /start",
        },
        
        "uz": {
            "welcome": "🤝 <b>ZARY & CO — ULGURJI BO'LIMI</b>\n\n"
                      "Do'konlar, butiklar va marketplace bilan ishlaymiz.\n"
                      "• Milliy brend\n"
                      "• Barqaror yetkazib berish\n"
                      "• Yuqori marja",
            
            "menu_hint": "📍 Asosiy menyu",
            
            "manager": f"📞 <b>Ulgurji bo'lim menejeri</b>\n\n"
                      f"Telefon: <code>{Config.PHONE}</code>\n"
                      f"Ish vaqti: Du-Ju 9:00-18:00",
            
            "channel": f"📣 <b>Bizning kanal</b>\n\n"
                      f"👉 https://t.me/{Config.CHANNEL}",
            
            "catalog": f"📸 <b>Dolzarb katalog</b>\n\n"
                      f"👉 https://t.me/{Config.CHANNEL}",
            
            "terms": "🧾 <b>Hamkorlik shartlari</b>\n\n"
                    "✅ Ish shakli: oldindan buyurtma / mavjud\n"
                    "✅ Minimal buyurtma: 20 donadan\n"
                    "✅ Yetkazib berish: O'zbekiston bo'ylab\n"
                    "✅ To'lov: o'tkazma / naqd",
            
            "why": "⭐ <b>Nega bizni tanlashadi</b>\n\n"
                  "🏆 Tajriba: 5+ yil\n"
                  "🏭 Ishlab chiqarish: Toshkentdagi o'zimizniki\n"
                  "📦 Assortiment: 500+ model\n"
                  "🚚 Logistika: butun mamlakat bo'ylab 2-3 kun",
            
            "min_order": "📦 <b>Minimal buyurtma</b>\n\n"
                        "• Ulgurji: 20 donadan\n"
                        "• Katta ulgurji: 100 donadan\n"
                        "• Dropshipping: alohida\n\n"
                        "Shaxsiy hisob-kitob xohlaysizmi?",
            
            "min_cta": "✅ Ariza qoldirish",
            
            "form_role": "👤 <b>Siz kimsiz?</b>",
            "form_product": "👕 <b>Nima buyurtma qilmoqchisiz?</b>",
            "form_qty": "📊 <b>Buyurtma hajmi?</b>",
            "form_city": "📍 <b>Yetkazib berish shahri?</b>",
            "form_phone": "📱 <b>Aloqa telefoni</b>\n\n"
                         "«📲 Kontakt yuborish» tugmasini bosing:",
            
            "bad_phone": "❌ <b>Noto'g'ri raqam</b>\n\n"
                        "Misol: +998901234567",
            
            "thanks": lambda lead_id: 
                     f"✅ <b>Ariza #{lead_id} qabul qilindi!</b>\n\n"
                     f"Menejer 15 daqiqa ichida bog'lanadi.\n\n"
                     f"📣 https://t.me/{Config.CHANNEL}\n"
                     f"📞 {Config.PHONE}",
            
            "cancelled": "❌ Bekor qilindi. Menyuga qaytish...",
            
            "admin_only": "⛔ Faqat admin uchun.",
            "admin_menu": "🛠 <b>Boshqaruv paneli</b>",
            "admin_last": "📋 <b>Oxirgi arizalar</b>\n\n",
            "admin_empty": "📝 Hozircha arizalar yo'q.",
            "admin_export_ok": "✅ Excel tayyor.",
            "admin_export_fail": "❌ Fayl yaratishda xatolik.",
            "admin_status_updated": "✅ Status yangilandi.",
            "admin_status_bad": "❌ Noto'g'ri buyruq.\n\n"
                               "Foydalanish: /status ID status\n\n"
                               "Statuslar: new, work, paid, shipped, closed",
            
            "monthly_report_subject": lambda p: f"📊 Hisobot: {p}",
            "monthly_report_intro": lambda s: 
                f"<b>📊 OYLIK HISOBOT — {s['period']}</b>\n\n"
                f"📅 Davr: {s['start']} — {s['end']}\n"
                f"📋 Jami arizalar: <b>{s['total']}</b>\n"
                f"👥 Unikal mijozlar: <b>{s['unique_clients']}</b>\n\n"
                f"<b>Statuslar:</b>\n"
                f"🆕 Yangi: {s['new_count']}\n"
                f"🔧 Ishlanmoqda: {s['work_count']}\n"
                f"💰 To'langan: {s['paid_count']}\n"
                f"🚚 Yuborilgan: {s['shipped_count']}\n"
                f"✅ Yopilgan: {s['closed_count']}",
            
            "error": "⚠️ Xatolik. Keyinroq urinib ko'ring.",
            "unknown": "🤔 Quyidagi menyudan foydalaning yoki /start",
        }
    }
    
    text = texts.get(lang, texts["ru"]).get(key, key)
    if callable(text):
        return text(**kwargs)
    return text

# =========================
# KEYBOARDS
# =========================
class Keyboards:
    @staticmethod
    def lang() -> ReplyKeyboardMarkup:
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="🇷🇺 Русский"), KeyboardButton(text="🇺🇿 O'zbekcha")]
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        )
    
    @staticmethod
    def main(lang: str, is_admin: bool) -> ReplyKeyboardMarkup:
        btn = lambda k: BTN[lang].get(k, k)
        rows = [
            [KeyboardButton(text=btn("catalog")), KeyboardButton(text=btn("terms"))],
            [KeyboardButton(text=btn("why")), KeyboardButton(text=btn("min"))],
            [KeyboardButton(text=btn("leave"))],
            [KeyboardButton(text=btn("manager")), KeyboardButton(text=btn("channel"))],
            [KeyboardButton(text=btn("lang"))],
        ]
        if is_admin:
            rows.append([KeyboardButton(text=btn("admin"))])
        return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)
    
    @staticmethod
    def form_role(lang: str) -> ReplyKeyboardMarkup:
        roles = {
            "uz": [["🏬 Butik", "🏪 Do'kon"], ["📱 Marketplace", "🌐 Boshqa"]],
            "ru": [["🏬 Бутик", "🏪 Магазин"], ["📱 Маркетплейс", "🌐 Другое"]]
        }
        r = roles.get(lang, roles["ru"])
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text=r[0][0]), KeyboardButton(text=r[0][1])],
                [KeyboardButton(text=r[1][0]), KeyboardButton(text=r[1][1])],
                [KeyboardButton(text=BTN[lang]["cancel"])]
            ],
            resize_keyboard=True
        )
    
    @staticmethod
    def form_product(lang: str) -> ReplyKeyboardMarkup:
        products = {
            "uz": [["👕 Kiyim", "👖 Shim"], ["🎒 Aksessuar", "👔 Boshqa"]],
            "ru": [["👕 Одежда", "👖 Брюки"], ["🎒 Аксессуары", "👔 Другое"]]
        }
        p = products.get(lang, products["ru"])
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text=p[0][0]), KeyboardButton(text=p[0][1])],
                [KeyboardButton(text=p[1][0]), KeyboardButton(text=p[1][1])],
                [KeyboardButton(text=BTN[lang]["cancel"])]
            ],
            resize_keyboard=True
        )
    
    @staticmethod
    def form_qty(lang: str) -> ReplyKeyboardMarkup:
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="20–50"), KeyboardButton(text="50–100")],
                [KeyboardButton(text="100–300"), KeyboardButton(text="300+")],
                [KeyboardButton(text=BTN[lang]["cancel"])]
            ],
            resize_keyboard=True
        )
    
    @staticmethod
    def form_phone(lang: str) -> ReplyKeyboardMarkup:
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text=BTN[lang]["contact"], request_contact=True)],
                [KeyboardButton(text=BTN[lang]["cancel"])]
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        )
    
    @staticmethod
    def min_cta(lang: str) -> ReplyKeyboardMarkup:
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text=t("min_cta", lang))],
                [KeyboardButton(text=BTN[lang]["cancel"])]
            ],
            resize_keyboard=True
        )
    
    @staticmethod
    def admin(lang: str) -> ReplyKeyboardMarkup:
        btn = lambda k: BTN[lang].get(k, k)
        return ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="📋 " + ("Последние" if lang == "ru" else "Oxirgi")), 
                 KeyboardButton(text="📊 " + ("Статистика" if lang == "ru" else "Statistika"))],
                [KeyboardButton(text="📤 Excel"), KeyboardButton(text="ℹ️ Status")],
                [KeyboardButton(text="⬅️ " + ("Назад" if lang == "ru" else "Orqaga"))]
            ],
            resize_keyboard=True
        )

BTN = {
    "ru": {
        "catalog": "📦 Каталог", "terms": "🧾 Условия", "why": "⭐ Почему мы",
        "min": "📦 Мин. заказ", "leave": "🤝 Заявка",
        "manager": "📞 Менеджер", "channel": "📣 Канал", "lang": "🌐 Язык",
        "admin": "🛠 Админ", "cancel": "❌ Отмена", "contact": "📲 Отправить контакт",
        "back": "Назад"
    },
    "uz": {
        "catalog": "📦 Katalog", "terms": "🧾 Shartlar", "why": "⭐ Nega biz",
        "min": "📦 Min. buyurtma", "leave": "🤝 Ariza",
        "manager": "📞 Menejer", "channel": "📣 Kanal", "lang": "🌐 Til",
        "admin": "🛠 Admin", "cancel": "❌ Bekor qilish", "contact": "📲 Kontakt yuborish",
        "back": "Orqaga"
    }
}

# =========================
# HELPERS
# =========================
def auto_lang(code: Optional[str]) -> str:
    return "uz" if (code or "").lower().startswith("uz") else "ru"

async def get_user_lang(message: Message) -> str:
    stored = await db.get_lang(message.from_user.id)
    if stored:
        return stored
    lang = auto_lang(message.from_user.language_code)
    await db.set_lang(message.from_user.id, lang)
    return lang

def is_admin(user_id: int) -> bool:
    return user_id == Config.MANAGER_ID

def normalize_phone(raw: str) -> str:
    if not raw:
        return ""
    s = re.sub(r"[^\d+]", "", raw.strip())
    if s.startswith("998") and not s.startswith("+998"):
        s = "+" + s
    elif s.startswith("9") and len(s) == 9:
        s = "+998" + s
    elif s.startswith("9") and len(s) == 12:
        s = "+998" + s[3:]
    return s

def is_valid_phone(phone: str) -> bool:
    p = normalize_phone(phone)
    if not p.startswith("+998"):
        return False
    digits = re.sub(r"\D", "", p)
    return len(digits) == 12 and digits[3:4] in "913"

# =========================
# FSM STATES
# =========================
class Form(StatesGroup):
    role = State()
    product = State()
    qty = State()
    city = State()
    phone = State()

# =========================
# HANDLERS
# =========================

@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await db.log_activity(message.from_user.id, "start", f"lang: {lang}")
    await message.answer(t("welcome", lang))
    await message.answer(t("menu_hint", lang), reply_markup=Keyboards.main(lang, is_admin(message.from_user.id)))

@dp.message(F.text.in_(["🇷🇺 Русский", "🇺🇿 O'zbekcha"]))
async def set_lang(message: Message, state: FSMContext):
    await state.clear()
    lang = "ru" if "Русский" in message.text else "uz"
    await db.set_lang(message.from_user.id, lang)
    await db.log_activity(message.from_user.id, "set_lang", lang)
    await message.answer(t("welcome", lang))
    await message.answer(t("menu_hint", lang), reply_markup=Keyboards.main(lang, is_admin(message.from_user.id)))

@dp.message(lambda m: m.text in {BTN["ru"]["lang"], BTN["uz"]["lang"]})
async def change_lang(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await message.answer(t("choose_lang", lang), reply_markup=Keyboards.lang())

# Menu handlers
@dp.message(lambda m: m.text in {BTN["ru"]["manager"], BTN["uz"]["manager"]})
async def menu_manager(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await db.log_activity(message.from_user.id, "view_manager")
    await message.answer(t("manager", lang), reply_markup=Keyboards.main(lang, is_admin(message.from_user.id)))

@dp.message(lambda m: m.text in {BTN["ru"]["channel"], BTN["uz"]["channel"]})
async def menu_channel(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await db.log_activity(message.from_user.id, "view_channel")
    await message.answer(t("channel", lang), reply_markup=Keyboards.main(lang, is_admin(message.from_user.id)))

@dp.message(lambda m: m.text in {BTN["ru"]["catalog"], BTN["uz"]["catalog"]})
async def menu_catalog(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await db.log_activity(message.from_user.id, "view_catalog")
    await message.answer(t("catalog", lang), reply_markup=Keyboards.main(lang, is_admin(message.from_user.id)))

@dp.message(lambda m: m.text in {BTN["ru"]["terms"], BTN["uz"]["terms"]})
async def menu_terms(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await db.log_activity(message.from_user.id, "view_terms")
    await message.answer(t("terms", lang), reply_markup=Keyboards.main(lang, is_admin(message.from_user.id)))

@dp.message(lambda m: m.text in {BTN["ru"]["why"], BTN["uz"]["why"]})
async def menu_why(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await db.log_activity(message.from_user.id, "view_why")
    await message.answer(t("why", lang), reply_markup=Keyboards.main(lang, is_admin(message.from_user.id)))

@dp.message(lambda m: m.text in {BTN["ru"]["min"], BTN["uz"]["min"]})
async def menu_min(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await db.log_activity(message.from_user.id, "view_min_order")
    await message.answer(t("min_order", lang), reply_markup=Keyboards.min_cta(lang))

# Form handlers
@dp.message(lambda m: m.text in {BTN["ru"]["leave"], BTN["uz"]["leave"], t("min_cta", "ru"), t("min_cta", "uz")})
async def form_start(message: Message, state: FSMContext):
    lang = await get_user_lang(message)
    await state.set_state(Form.role)
    await db.log_activity(message.from_user.id, "start_form")
    await message.answer(t("form_role", lang), reply_markup=Keyboards.form_role(lang))

@dp.message(Form.role)
async def form_role(message: Message, state: FSMContext):
    lang = await get_user_lang(message)
    text = (message.text or "").strip()
    if text in {BTN["ru"]["cancel"], BTN["uz"]["cancel"]}:
        await cancel_handler(message, state)
        return
    await state.update_data(role=text)
    await state.set_state(Form.product)
    await message.answer(t("form_product", lang), reply_markup=Keyboards.form_product(lang))

@dp.message(Form.product)
async def form_product(message: Message, state: FSMContext):
    lang = await get_user_lang(message)
    text = (message.text or "").strip()
    if text in {BTN["ru"]["cancel"], BTN["uz"]["cancel"]}:
        await cancel_handler(message, state)
        return
    await state.update_data(product=text)
    await state.set_state(Form.qty)
    await message.answer(t("form_qty", lang), reply_markup=Keyboards.form_qty(lang))

@dp.message(Form.qty)
async def form_qty(message: Message, state: FSMContext):
    lang = await get_user_lang(message)
    text = (message.text or "").strip()
    if text in {BTN["ru"]["cancel"], BTN["uz"]["cancel"]}:
        await cancel_handler(message, state)
        return
    await state.update_data(qty=text)
    await state.set_state(Form.city)
    await message.answer(t("form_city", lang), reply_markup=ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=BTN[lang]["cancel"])]],
        resize_keyboard=True
    ))

@dp.message(Form.city)
async def form_city(message: Message, state: FSMContext):
    lang = await get_user_lang(message)
    text = (message.text or "").strip()
    if text in {BTN["ru"]["cancel"], BTN["uz"]["cancel"]}:
        await cancel_handler(message, state)
        return
    await state.update_data(city=text)
    await state.set_state(Form.phone)
    await message.answer(t("form_phone", lang), reply_markup=Keyboards.form_phone(lang))

@dp.message(Form.phone)
async def form_phone(message: Message, state: FSMContext):
    lang = await get_user_lang(message)
    
    if (message.text or "").strip() in {BTN["ru"]["cancel"], BTN["uz"]["cancel"]}:
        await cancel_handler(message, state)
        return
    
    raw_phone = ""
    if message.contact:
        raw_phone = message.contact.phone_number
    else:
        raw_phone = (message.text or "").strip()
    
    phone = normalize_phone(raw_phone)
    if not is_valid_phone(phone):
        await message.answer(t("bad_phone", lang))
        return
    
    data = await state.get_data()
    user = message.from_user
    
    lead = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "user_id": user.id,
        "username": user.username,
        "full_name": user.full_name,
        "lang": lang,
        "role": data.get("role", "-"),
        "product": data.get("product", "-"),
        "qty": data.get("qty", "-"),
        "city": data.get("city", "-"),
        "phone": phone,
    }
    
    try:
        lead_id = await db.add_lead(lead)
        await db.log_activity(user.id, "lead_created", f"lead_id: {lead_id}")
        await notify_manager(lead, lead_id, lang)
        await message.answer(t("thanks", lang, lead_id=lead_id), 
                           reply_markup=Keyboards.main(lang, is_admin(user.id)))
    except Exception as e:
        logger.exception("Failed to save lead")
        await message.answer(t("error", lang))
    
    await state.clear()

async def notify_manager(lead: dict, lead_id: int, client_lang: str):
    lang_label = "🇷🇺 RU" if client_lang == "ru" else "🇺🇿 UZ"
    msg = (
        f"🔔 <b>Новая заявка #{lead_id}</b> {lang_label}\n\n"
        f"👤 {lead['full_name']}\n"
        f"📱 <code>{lead['phone']}</code>\n"
        f"🏢 {lead['role']} | {lead['product']} | {lead['qty']}\n"
        f"📍 {lead['city']}\n"
        f"⏰ {lead['created_at']}\n\n"
        f"👤 @{lead['username'] or 'нет'}\n"
        f"🆔 <code>{lead['user_id']}</code>\n"
        f"📋 /status {lead_id} work"
    )
    try:
        await bot.send_message(Config.MANAGER_ID, msg)
        await db.update_notification_status(lead_id, True)
    except TelegramAPIError as e:
        logger.error(f"Failed to notify manager: {e}")
        await db.log_activity(lead['user_id'], "notify_failed", str(e))

async def cancel_handler(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await message.answer(t("cancelled", lang), 
                       reply_markup=Keyboards.main(lang, is_admin(message.from_user.id)))

@dp.message(lambda m: m.text in {BTN["ru"]["cancel"], BTN["uz"]["cancel"]})
async def cmd_cancel(message: Message, state: FSMContext):
    await cancel_handler(message, state)

# =========================
# ADMIN HANDLERS
# =========================
@dp.message(lambda m: m.text in {BTN["ru"]["admin"], BTN["uz"]["admin"]})
async def admin_menu(message: Message, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id):
        lang = await get_user_lang(message)
        await message.answer(t("admin_only", lang))
        return
    lang = await get_user_lang(message)
    await message.answer(t("admin_menu", lang), reply_markup=Keyboards.admin(lang))

@dp.message(lambda m: m.text.startswith("📋"))
async def admin_last(message: Message, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id):
        return
    lang = await get_user_lang(message)
    rows = await db.get_last_leads(20)
    if not rows:
        await message.answer(t("admin_empty", lang), reply_markup=Keyboards.admin(lang))
        return
    
    lines = [t("admin_last", lang)]
    for r in rows:
        status_emoji = {"new": "🆕", "work": "🔧", "paid": "💰", "shipped": "🚚", "closed": "✅"}.get(r["status"], "❓")
        lines.append(
            f"\n<b>#{r['id']}</b> {status_emoji} <code>{r['status']}</code>\n"
            f"📅 {r['created_at'][:16]} | {r['role']} | {r['product']}\n"
            f"📍 {r['city']} | ☎️ {r['phone']}\n"
            f"{'✓' if r['manager_notified'] else '✗'} | {r['user_id']}\n"
            f"──────────────"
        )
    await message.answer("\n".join(lines), reply_markup=Keyboards.admin(lang))

@dp.message(lambda m: m.text.startswith("📊"))
async def admin_stats(message: Message, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id):
        return
    lang = await get_user_lang(message)
    stats = await db.get_stats()
    await message.answer(t("stats", lang, s=stats) if "stats" in str(t("stats", lang, s=stats)) else f"📊 Статистика:\n\n• Всего: {stats['total_leads']}\n• Новых: {stats['new_leads']}\n• Клиентов: {stats['unique_users']}", 
                       reply_markup=Keyboards.admin(lang))

@dp.message(Command("status"))
async def admin_set_status(message: Message, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id):
        lang = await get_user_lang(message)
        await message.answer(t("admin_only", lang))
        return
    
    lang = await get_user_lang(message)
    parts = (message.text or "").split()
    if len(parts) != 3 or not parts[1].isdigit():
        await message.answer(t("admin_status_bad", lang), reply_markup=Keyboards.admin(lang))
        return
    
    lead_id = int(parts[1])
    status = parts[2].lower()
    if status not in {"new", "work", "paid", "shipped", "closed"}:
        await message.answer(t("admin_status_bad", lang), reply_markup=Keyboards.admin(lang))
        return
    
    success = await db.update_status(lead_id, status)
    if success:
        await message.answer(t("admin_status_updated", lang), reply_markup=Keyboards.admin(lang))
        await db.log_activity(message.from_user.id, "status_update", f"lead {lead_id} -> {status}")
    else:
        await message.answer(f"❌ Заявка #{lead_id} не найдена", reply_markup=Keyboards.admin(lang))

@dp.message(lambda m: m.text == "📤 Excel")
async def admin_export(message: Message, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id):
        return
    
    lang = await get_user_lang(message)
    try:
        rows = await db.get_all_leads()
        if not rows:
            await message.answer(t("admin_empty", lang), reply_markup=Keyboards.admin(lang))
            return
        
        Config.EXPORTS_DIR.mkdir(exist_ok=True)
        filename = await create_excel(rows, Config.EXPORTS_DIR / f"leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        await message.answer(t("admin_export_ok", lang), reply_markup=Keyboards.admin(lang))
        await bot.send_document(
            message.from_user.id,
            FSInputFile(str(filename)),
            caption=f"📤 Экспорт от {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )
        await db.log_activity(message.from_user.id, "export_excel", str(filename))
    except Exception as e:
        logger.exception("Excel export failed")
        await message.answer(t("admin_export_fail", lang), reply_markup=Keyboards.admin(lang))

@dp.message(lambda m: m.text.startswith("⬅️"))
async def admin_back(message: Message, state: FSMContext):
    await state.clear()
    lang = await get_user_lang(message)
    await message.answer(t("menu_hint", lang), reply_markup=Keyboards.main(lang, is_admin(message.from_user.id)))

# =========================
# EXCEL CREATION (универсальная функция)
# =========================
async def create_excel(rows: List[aiosqlite.Row], filepath: Path, title: str = "Leads") -> Path:
    """Создать красивый Excel файл"""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    # Заголовки с стилем
    headers = ["ID", "Дата", "Клиент", "Username", "Язык", "Тип", "Товар", 
              "Кол-во", "Город", "Телефон", "Статус", "Уведомлен"]
    
    ws.append(headers)
    
    # Стили для заголовка
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    for r in rows:
        ws.append([
            r["id"], r["created_at"], r["full_name"], r["username"], r["lang"],
            r["role"], r["product"], r["qty"], r["city"], r["phone"],
            r["status"], "Да" if r["manager_notified"] else "Нет"
        ])
    
    # Автоширина
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    wb.save(filepath)
    return filepath

# =========================
# MONTHLY REPORT (автоотчет в конце месяца)
# =========================
async def send_monthly_report():
    """Отправить месячный отчет админу"""
    now = datetime.now()
    year, month = now.year, now.month
    
    # Проверяем не отправляли ли уже
    if await db.is_report_sent(year, month):
        logger.info(f"Report for {month}.{year} already sent")
        return
    
    # Получаем статистику
    stats = await db.get_monthly_stats(year, month)
    
    if stats['total'] == 0:
        logger.info(f"No leads for {month}.{year}, skipping report")
        return
    
    # Создаем Excel
    start_date = f"{year}-{month:02d}-01"
    last_day = monthrange(year, month)[1]
    end_date = f"{year}-{month:02d}-{last_day} 23:59:59"
    
    rows = await db.get_leads_by_date_range(start_date, end_date)
    
    Config.REPORTS_DIR.mkdir(exist_ok=True)
    filename = Config.REPORTS_DIR / f"monthly_report_{year}_{month:02d}.xlsx"
    
    await create_excel(rows, filename, f"Report_{month:02d}_{year}")
    
    # Отправляем админу
    try:
        # Текстовая сводка
        intro = t("monthly_report_intro", "ru", s=stats)
        await bot.send_message(Config.MANAGER_ID, intro)
        
        # Excel файл
        await bot.send_document(
            Config.MANAGER_ID,
            FSInputFile(str(filename)),
            caption=f"📊 Полный отчет за {stats['period']}\n\n"
                   f"Всего заявок: {stats['total']}\n"
                   f"Файл: {filename.name}"
        )
        
        # Отмечаем как отправленный
        await db.mark_report_sent(year, month, str(filename), stats['total'])
        logger.info(f"Monthly report for {month}.{year} sent successfully")
        
    except TelegramAPIError as e:
        logger.error(f"Failed to send monthly report: {e}")

# =========================
# CLEANUP & BACKUP
# =========================
async def cleanup_old_files():
    """Очистка старых файлов"""
    try:
        cutoff = datetime.now() - timedelta(days=Config.MAX_EXPORT_AGE_DAYS)
        count = 0
        for file in Config.EXPORTS_DIR.glob("*.xlsx"):
            if datetime.fromtimestamp(file.stat().st_mtime) < cutoff:
                file.unlink()
                count += 1
        if count > 0:
            logger.info(f"Cleaned up {count} old export files")
    except Exception as e:
        logger.error(f"Cleanup error: {e}")

async def backup_database():
    """Резервное копирование"""
    try:
        Config.BACKUP_DIR.mkdir(exist_ok=True)
        backup_path = Config.BACKUP_DIR / f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        import shutil
        shutil.copy(Config.DB_PATH, backup_path)
        
        # Удаляем старые бэкапы
        backups = sorted(Config.BACKUP_DIR.glob("*.db"), key=lambda p: p.stat().st_mtime)
        for old in backups[:-Config.BACKUP_KEEP_COUNT]:
            old.unlink()
        
        logger.info(f"Database backed up to {backup_path}")
    except Exception as e:
        logger.error(f"Backup error: {e}")

# =========================
# WEB SERVER
# =========================
async def health_check(request):
    return web.Response(text="OK", status=200)

async def start_web_server():
    app = web.Application()
    app.router.add_get("/", health_check)
    app.router.add_get("/health", health_check)
    
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", Config.PORT)
    await site.start()
    logger.info(f"Web server started on port {Config.PORT}")

# =========================
# MAIN
# =========================
async def main():
    # Подключаем БД
    await db.connect()
    
    # Очищаем при старте
    await cleanup_old_files()
    
    # Настраиваем планировщик
    scheduler = AsyncIOScheduler()
    
    # Ежедневные задачи
    scheduler.add_job(cleanup_old_files, "cron", hour=3, minute=0)
    scheduler.add_job(backup_database, "cron", hour=2, minute=0)
    
    # ⭐ ВАЖНО: Автоотчет в последний день месяца в 23:00
    scheduler.add_job(send_monthly_report, "cron", day="last", hour=23, minute=0)
    
    # Также проверяем при старте (если бот был выключен в последний день)
    scheduler.add_job(send_monthly_report, "date", run_date=datetime.now() + timedelta(seconds=30))
    
    scheduler.start()
    
    # Удаляем webhook
    await bot.delete_webhook(drop_pending_updates=True)
    
    # Запускаем
    await asyncio.gather(
        start_web_server(),
        dp.start_polling(bot, skip_updates=True)
    )

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logger.info("Bot stopped")
    except Exception as e:
        logger.critical(f"Fatal error: {e}")
        raise
